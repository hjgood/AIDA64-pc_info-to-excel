# -*- coding: utf-8 -*-
import _locale

_locale._getdefaultlocale = (lambda *args: ['zh_CN', 'utf8'])
import sys
import os
import openpyxl
import datetime
from chardet.universaldetector import UniversalDetector

# 定义变量
file_dir = './inv/'
book = 'IT客户端清单.xlsx'


def get_encode_info(file):
    with open(file_dir + file, 'rb') as f:
        detector = UniversalDetector()
        for line in f.readlines():
            detector.feed(line)
            if detector.done:
                break
        detector.close()
        return detector.result['encoding']


def read_file(file):
    with open(file_dir + file, 'rb') as f:
        return f.read()


def write_file(content, file):
    with open(file_dir + file, 'wb') as f:
        f.write(content)


def convert_encode2utf8(file, original_encode, des_encode):
    file_content = read_file(file)
    file_decode = file_content.decode(original_encode, 'ignore')
    file_encode = file_decode.encode(des_encode)
    write_file(file_encode, file)


def yeah():  # 读年份
    yeah = datetime.datetime.now().year
    return yeah


def listdir():  # 列目录文件
    for root, dirs, files in os.walk(file_dir, topdown=False):
        list = files
    return list


def openexecl(book):  # 打开IT清单
    wb = openpyxl.load_workbook(book)
    return wb


def txtname(name):  # 取文件文件名去除扩展名用作第一字段
    filen_name = name.split('.')[0]
    return filen_name


def txtinfo():  # 抓取文本相关记录内容
    all = []

    for name in listdir():
        filename = (file_dir + name)
        filen_name = name.split('.')[0]
        f = open(filename, 'r', encoding='utf8')
        source = f.readlines()
        f.close()
        hddx = []
        time = ''
        sn = ''
        cpu = ''
        vga = ''
        os = ''
        conut = []
        info = []
        mac = ''
        main = ''
        mem = ''
        d3 = ''
        hdd = ''
        mac = ''
        all1 = []
        all.append(filen_name)
        for x in source:
            a = x.split()
            if a != []:
                if a[0] == ('序列号'):
                    li = a[0]
                    if li not in conut:  # 记录唯一
                        c = ' '.join(a[1:])
                        sn = c
                        conut.append(li)
            if a != []:
                if a[0] == ('操作系统'):
                    li = a[0]
                    if li not in conut:  # 记录唯一
                        c = ' '.join(a[1:])
                        os = c
                        conut.append(li)
            if a != []:
                if a[0] == ('处理器名称'):
                    li = a[0]
                    if li not in conut:  # 记录唯一
                        c = ' '.join(a[1:])
                        cpu = c
                        conut.append(li)
            if a != []:
                if a[0] == ('显示器'):
                    li = a[0]
                    if li not in conut:  # 记录唯一
                        c = ' '.join(a[1:])
                        d3 = c
                        conut.append(li)
            if a != []:
                if a[0] == ('主板名称'):
                    li = a[0]
                    if li not in conut:  # 记录唯一
                        c = ' '.join(a[1:])
                        main = c
                        conut.append(li)
            if a != []:
                if a[0] == ('系统内存'):
                    li = a[0]
                    if li not in conut:  # 记录唯一
                        c = ' '.join(a[1:])
                        mem = c
                        conut.append(li)
            if a != []:
                if a[0] == ('总大小'):
                    li = a[0]
                    if li not in conut:  # 记录唯一
                        c = ' '.join(a[1:])
                        hdd = c
                        conut.append(li)
            if a != []:
                if a[0] == ('显示适配器'):
                    li = a[0]
                    if li not in conut:  # 记录唯一
                        c = ' '.join(a[1:])
                        vga = c
                        conut.append(li)

            if a != []:
                if a[0] in ('主'):
                    if a[1] == ('MAC'):
                        if a[1] not in conut:
                            c = a[3]
                            mac = c
                            conut.append(a[3])
            if a != []:  # 多硬盘
                if a[0] == ('硬盘驱动器'):
                    # print (a)
                    if a[1] not in ('SMART 状态 OK)'):
                        if a[-1] != ('USB)'):
                            if a[1] not in ('Generic- Multi-Card USB Device'):
                                d = ' '.join(a[1:])
                                hddx.append(d)
            if a !=[]:
                if a[0] == ('发布日期'):
                    li = a[0]
                    if li not in conut:  # 记录唯一
                        c = ' '.join(a[1:])
                        time = c
                        conut.append(li)
        hddx_no = len(hddx)
        if hddx_no != 2:
            hddx.append(' ')
        #all1 = (os + cpu + main + mem + d3 + vga + hdd + mac + time + sn)
        #all1 = all1.append(hddx)
        all.append(os)
        all.append(cpu)
        all.append(main)
        all.append(mem)
        all.append(d3)
        all.append(vga)
        all.append(hdd)
        all.append(mac)
        all.append(time)
        all.append(sn)
        all = all + hddx
    return all


def writeexcel(txtinfo):
    row1 = 5
    column1 = 4
    no = 13
    wb = openpyxl.load_workbook('IT客户端清单.xlsx')
    ws1 = wb['库存列表']
    for one in txtinfo:
        ws1.cell(row=row1, column=column1).value = one
        if no != 0:
            print(row1, column1, one)  # ,no)
            no = no - 1
            column1 = column1 + 1
        if no == 0:
            row1 = row1 + 1
            no = 13
            column1 = 4
        wb.save('IT客户端清单.xlsx')


def format():  # 整理单元格内容
    sumlist = len(listdir())
    row1 = 5
    column1 = 4
    wb = openpyxl.load_workbook('IT客户端清单.xlsx')
    ws1 = wb['库存列表']
    for nor in range(sumlist):
        os = ws1.cell(row1, column1 + 1).value
        cpu = ws1.cell(row1, column1 + 2).value
        main = ws1.cell(row1, column1 + 3).value
        mem = ws1.cell(row1, column1 + 4).value
        d3 = ws1.cell(row1, column1 + 5).value
        vga = ws1.cell(row1, column1 + 6).value
        hdd = ws1.cell(row1, column1 + 7).value
        mac = ws1.cell(row1, column1 + 8).value
        yeahs = ws1.cell(row1, column1 + 9).value
        yeahs = yeahs.split('/')[2]
        yeahs = int(yeah()) - int(yeahs)
        sn = ws1.cell(row1, column1 + 10).value
        hdd1 = ws1.cell(row1, column1 + 11).value
        hdd2 = ws1.cell(row1, column1 + 12).value
        ws1.cell(row1, column1 + 1).value = sn
        ws1.cell(row1, column1 + 2).value = yeahs
        ws1.cell(row1, column1 + 3).value = os
        ws1.cell(row1, column1 + 4).value = d3
        ws1.cell(row1, column1 + 5).value = cpu
        ws1.cell(row1, column1 + 6).value = main
        ws1.cell(row1, column1 + 7).value = mem
        ws1.cell(row1, column1 + 8).value = vga
        ws1.cell(row1, column1 + 9).value = hdd
        ws1.cell(row1, column1 + 10).value = mac
        ws1.cell(row1, column1 + 11).value = str(hdd1) + '\n' + str(hdd2)
        ws1.cell(row1, column1 + 12).value = ' '
        row1 = row1 + 1
        wb.save('IT客户端清单.xlsx')


if __name__ == '__main__':
    for filename in listdir():
        file_content = read_file(filename)
        encode_info = get_encode_info(filename)
        if encode_info != 'utf-8':
            convert_encode2utf8(filename, encode_info, 'utf-8')
        encode_info = get_encode_info(filename)
        # print(filename + ' ' + encode_info) #显示文件名与转换后编码

    try:
        openexecl(book).save('IT客户端清单.xlsx')
    except:
        input('-----请先关闭IT客户端清单.xlsx-----')
        sys.exit(1)
    txtinfo = txtinfo()
    writeexcel(txtinfo)
    format()
    sumlist = len(listdir())
    input('！！！共计导入 ' + str(sumlist) + ' 台主机信息！！！')
